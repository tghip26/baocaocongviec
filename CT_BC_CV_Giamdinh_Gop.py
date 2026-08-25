import os
import re
import warnings
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==========================================================
# ẨN CẢNH BÁO OPENPYXL
# ==========================================================

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style"
)


# ==========================================================
# TÊN FILE XUẤT
# ==========================================================

TEN_FILE_GIAM_DINH = (
    "DANH SÁCH THEO DÕI GIÁM ĐỊNH BẢO HIỂM.xlsx"
)

TEN_FILE_CNTT = (
    "TỔNG HỢP CÔNG VIỆC P.CNTT.xlsx"
)


# ==========================================================
# DANH SÁCH NGƯỜI DÙNG
# ==========================================================

ALLOWED_USERS = {
    "PHCN",
    "NGOẠI TH",
    "XNTT",
    "NHH",
    "TTTM",
    "NGOẠI CT",
    "DA LIỄU",
    "PTGMHS",
    "NỘI TH",
    "PHỤ SẢN",
    "MẮT",
    "TTBVSK",
    "DƯỢC",
    "THẬN",
    "TRUYỀN NHIỄM",
    "TCKT",
    "NGOẠI UB",
    "NGOẠI XẠ TRỊ",
    "RHM",
    "KB",
    "CXK",
    "CSGN",
    "NGOẠI TN",
    "HHLS",
    "TMH",
    "HSTC",
    "HTL",
    "NGOẠI TKLN",
    "NHI",
    "ĐÔNG Y",
    "CẤP CỨU",
    "LKTK"
}


# ==========================================================
# HÀM CHUẨN HÓA TEXT
# ==========================================================

def normalize_text(value):

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\r", " ")
    text = text.replace("\n", " ")

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip().upper()


# ==========================================================
# SAO CHÉP ĐỊNH DẠNG
# ==========================================================

def copy_style(source, target):

    if source.has_style:
        target._style = copy(
            source._style
        )

    target.font = copy(
        source.font
    )

    target.fill = copy(
        source.fill
    )

    target.border = copy(
        source.border
    )

    target.alignment = copy(
        source.alignment
    )

    target.protection = copy(
        source.protection
    )

    target.number_format = (
        source.number_format
    )


# ==========================================================
# CĂN GIỮA
# ==========================================================

def center_cell(
    cell,
    wrap=False
):

    alignment = copy(
        cell.alignment
    )

    alignment.horizontal = "center"
    alignment.vertical = "center"
    alignment.wrap_text = wrap

    cell.alignment = alignment


# ==========================================================
# ==========================================================
# CHỨC NĂNG 1
# BÁO CÁO CỔNG GIÁM ĐỊNH
# ==========================================================
# ==========================================================


def find_header_row(ws):

    for row in range(
        1,
        min(ws.max_row, 20) + 1
    ):

        values = []

        for col in range(
            1,
            ws.max_column + 1
        ):

            values.append(
                normalize_text(
                    ws.cell(
                        row,
                        col
                    ).value
                )
            )

        if (
            "STT" in values
            and
            "NGƯỜI DÙNG" in values
        ):

            return row

    raise Exception(
        "Không tìm thấy dòng tiêu đề STT / NGƯỜI DÙNG."
    )


def find_columns(
    ws,
    header_row
):

    user_col = None

    day_columns = {}

    for col in range(
        1,
        ws.max_column + 1
    ):

        value = ws.cell(
            header_row,
            col
        ).value

        text = normalize_text(
            value
        )

        if text == "NGƯỜI DÙNG":

            user_col = col

        if value is not None:

            value_text = str(
                value
            ).strip()

            if value_text.isdigit():

                day = int(
                    value_text
                )

                if 1 <= day <= 31:

                    day_columns[
                        day
                    ] = col

    if user_col is None:

        raise Exception(
            "Không tìm thấy cột NGƯỜI DÙNG."
        )

    if not day_columns:

        raise Exception(
            "Không tìm thấy cột ngày."
        )

    return (
        user_col,
        day_columns
    )


def choose_day_range():

    today = datetime.datetime.now().day

    if today <= 14:
        default_value = "01-14"
    else:
        default_value = "15-31"

    result = {
        "value": None
    }

    window = tk.Toplevel()

    window.title(
        "Chọn khoảng ngày"
    )

    window.geometry(
        "420x270"
    )

    window.resizable(
        False,
        False
    )

    tk.Label(
        window,
        text="CHỌN KHOẢNG NGÀY",
        font=(
            "Arial",
            13,
            "bold"
        )
    ).pack(
        pady=15
    )

    tk.Label(
        window,
        text=(
            f"Ngày hiện tại: "
            f"{today:02d}\n"
            f"Mặc định: "
            f"{default_value}"
        )
    ).pack(
        pady=5
    )

    selected = tk.StringVar(
        value=default_value
    )

    tk.Radiobutton(
        window,
        text="Ngày 01 → 14",
        variable=selected,
        value="01-14",
        font=(
            "Arial",
            11
        )
    ).pack(
        anchor="w",
        padx=80,
        pady=5
    )

    tk.Radiobutton(
        window,
        text="Ngày 15 → 31",
        variable=selected,
        value="15-31",
        font=(
            "Arial",
            11
        )
    ).pack(
        anchor="w",
        padx=80,
        pady=5
    )

    def confirm():

        result["value"] = (
            selected.get()
        )

        window.destroy()

    tk.Button(
        window,
        text="TIẾP TỤC",
        command=confirm,
        width=18,
        height=2
    ).pack(
        pady=15
    )

    window.grab_set()

    window.wait_window()

    return result["value"]


def process_giam_dinh(
    input_file,
    start_day,
    end_day
):

    # ------------------------------------------------------
    # ĐỌC FILE
    # ------------------------------------------------------

    source_wb = openpyxl.load_workbook(
        input_file
    )

    if (
        "Người dùng"
        not in source_wb.sheetnames
    ):

        raise Exception(
            "File không có sheet 'Người dùng'."
        )

    source = source_wb[
        "Người dùng"
    ]

    # ------------------------------------------------------
    # TÌM HEADER
    # ------------------------------------------------------

    header_row = find_header_row(
        source
    )

    user_col, day_columns = (
        find_columns(
            source,
            header_row
        )
    )

    # ------------------------------------------------------
    # LẤY NGÀY
    # ------------------------------------------------------

    selected_days = []

    for day in range(
        start_day,
        end_day + 1
    ):

        if day not in day_columns:

            raise Exception(
                f"Không tìm thấy cột ngày {day}."
            )

        selected_days.append(
            day
        )

    # ------------------------------------------------------
    # CỘT KẾT QUẢ
    #
    # A = STT
    # B = NGƯỜI DÙNG
    # C trở đi = NGÀY
    # ------------------------------------------------------

    source_columns = [
        1,
        2
    ]

    for day in selected_days:

        source_columns.append(
            day_columns[day]
        )

    result_last_col = len(
        source_columns
    )

    result_last_letter = (
        get_column_letter(
            result_last_col
        )
    )

    # ------------------------------------------------------
    # TẠO FILE MỚI
    # ------------------------------------------------------

    result_wb = Workbook()

    result = result_wb.active

    result.title = "Người dùng"

    # ------------------------------------------------------
    # COPY DÒNG 1-5
    # ------------------------------------------------------

    for row in range(
        1,
        6
    ):

        for new_col, old_col in enumerate(
            source_columns,
            start=1
        ):

            source_cell = source.cell(
                row,
                old_col
            )

            target_cell = result.cell(
                row,
                new_col
            )

            target_cell.value = (
                source_cell.value
            )

            copy_style(
                source_cell,
                target_cell
            )

    # ------------------------------------------------------
    # GỘP A1:P1 HOẶC A1:S1
    # ------------------------------------------------------

    title = source[
        "A1"
    ].value

    for col in range(
        2,
        result_last_col + 1
    ):

        result.cell(
            1,
            col
        ).value = None

    result[
        "A1"
    ] = title

    result.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=result_last_col
    )

    center_cell(
        result["A1"],
        wrap=False
    )

    result.row_dimensions[
        1
    ].height = 32

    # ------------------------------------------------------
    # GỘP A2:P2 HOẶC A2:S2
    # ------------------------------------------------------

    note = source[
        "A2"
    ].value

    for col in range(
        2,
        result_last_col + 1
    ):

        result.cell(
            2,
            col
        ).value = None

    result[
        "A2"
    ] = note

    result.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=result_last_col
    )

    # ------------------------------------------------------
    # GỘP A4:A5
    # ------------------------------------------------------

    result[
        "A5"
    ] = None

    result.merge_cells(
        "A4:A5"
    )

    center_cell(
        result["A4"],
        wrap=False
    )

    # ------------------------------------------------------
    # GỘP B4:B5
    # ------------------------------------------------------

    result[
        "B5"
    ] = None

    result.merge_cells(
        "B4:B5"
    )

    center_cell(
        result["B4"],
        wrap=False
    )

    # ------------------------------------------------------
    # LỌC + ĐÁNH STT LẠI
    # ------------------------------------------------------

    output_row = 6

    stt = 1

    for source_row in range(
        6,
        source.max_row + 1
    ):

        user_name = normalize_text(
            source.cell(
                source_row,
                user_col
            ).value
        )

        if (
            user_name
            not in ALLOWED_USERS
        ):
            continue

        for new_col, old_col in enumerate(
            source_columns,
            start=1
        ):

            source_cell = source.cell(
                source_row,
                old_col
            )

            target_cell = result.cell(
                output_row,
                new_col
            )

            # STT MỚI
            if new_col == 1:

                target_cell.value = stt

            else:

                target_cell.value = (
                    source_cell.value
                )

            copy_style(
                source_cell,
                target_cell
            )

        stt += 1

        output_row += 1

    # ------------------------------------------------------
    # KÍCH THƯỚC CỘT
    # ------------------------------------------------------

    for new_col, old_col in enumerate(
        source_columns,
        start=1
    ):

        old_letter = (
            get_column_letter(
                old_col
            )
        )

        new_letter = (
            get_column_letter(
                new_col
            )
        )

        width = source.column_dimensions[
            old_letter
        ].width

        if width:

            result.column_dimensions[
                new_letter
            ].width = width

    result.column_dimensions[
        "A"
    ].width = 8

    result.column_dimensions[
        "B"
    ].width = 25

    result.freeze_panes = "C6"

    result.auto_filter.ref = (
        f"A4:{result_last_letter}"
        f"{result.max_row}"
    )

    # ------------------------------------------------------
    # LƯU CÙNG THƯ MỤC FILE GỐC
    # ------------------------------------------------------

    folder = os.path.dirname(
        input_file
    )

    output_file = os.path.join(
        folder,
        TEN_FILE_GIAM_DINH
    )

    # Nếu file đã tồn tại thì hỏi ghi đè

    if os.path.exists(
        output_file
    ):

        answer = messagebox.askyesno(
            "File đã tồn tại",
            (
                f"File:\n"
                f"{TEN_FILE_GIAM_DINH}\n\n"
                "đã tồn tại.\n\n"
                "Bạn có muốn ghi đè không?"
            )
        )

        if not answer:
            return None

    result_wb.save(
        output_file
    )

    # ------------------------------------------------------
    # KIỂM TRA SAU KHI LƯU
    # ------------------------------------------------------

    check_wb = openpyxl.load_workbook(
        output_file
    )

    check = check_wb[
        "Người dùng"
    ]

    actual_merges = {
        str(x)
        for x in check.merged_cells.ranges
    }

    required_merges = {
        f"A1:{result_last_letter}1",
        f"A2:{result_last_letter}2",
        "A4:A5",
        "B4:B5"
    }

    missing = (
        required_merges
        - actual_merges
    )

    if missing:

        raise Exception(
            "Gộp ô chưa đúng:\n"
            + "\n".join(
                sorted(missing)
            )
        )

    # Kiểm tra STT

    if check["A6"].value != 1:

        raise Exception(
            f"A6 phải bằng 1 "
            f"nhưng đang là "
            f"{check['A6'].value}"
        )

    if check.max_row >= 7:

        if check["A7"].value != 2:

            raise Exception(
                f"A7 phải bằng 2 "
                f"nhưng đang là "
                f"{check['A7'].value}"
            )

    if check.max_row >= 8:

        if check["A8"].value != 3:

            raise Exception(
                f"A8 phải bằng 3 "
                f"nhưng đang là "
                f"{check['A8'].value}"
            )

    return output_file


# ==========================================================
# ==========================================================
# CHỨC NĂNG 2
# BÁO CÁO CÔNG VIỆC P.CNTT
# ==========================================================
# ==========================================================


def process_cntt(
    input_file,
    output_file
):

    # ------------------------------------------------------
    # ĐỌC FILE
    # ------------------------------------------------------

    wb_source = load_workbook(
        input_file,
        data_only=True
    )

    # ------------------------------------------------------
    # CHỌN SHEET
    # ------------------------------------------------------

    if len(
        wb_source.sheetnames
    ) == 1:

        sheet_name = (
            wb_source.sheetnames[0]
        )

    else:

        sheet_window = tk.Toplevel()

        sheet_window.title(
            "Chọn Sheet P.CNTT"
        )

        sheet_window.geometry(
            "450x180"
        )

        tk.Label(
            sheet_window,
            text="CHỌN SHEET CẦN XỬ LÝ",
            font=(
                "Arial",
                12,
                "bold"
            )
        ).pack(
            pady=15
        )

        combo = ttk.Combobox(
            sheet_window,
            values=wb_source.sheetnames,
            state="readonly",
            width=45
        )

        combo.pack(
            pady=5
        )

        combo.current(0)

        selected = {
            "value": None
        }

        def confirm_sheet():

            selected["value"] = (
                combo.get()
            )

            sheet_window.destroy()

        tk.Button(
            sheet_window,
            text="TIẾP TỤC",
            command=confirm_sheet,
            width=15
        ).pack(
            pady=15
        )

        sheet_window.grab_set()

        sheet_window.wait_window()

        sheet_name = selected[
            "value"
        ]

        if not sheet_name:

            return None

    ws_source = wb_source[
        sheet_name
    ]

    # ======================================================
    # CẤU TRÚC FILE GỐC
    #
    # Hàng 105: tên cột công việc
    # Hàng 106: dữ liệu khoa
    # Cột D: Khoa
    # Cột E:Y: Phần mềm
    # Cột AI: Sửa chữa khác
    # ======================================================

    software_columns = list(
        range(5, 26)
    )

    software_headers = []

    for col in software_columns:

        header = ws_source.cell(
            105,
            col
        ).value

        if header is None:

            header = ""

        software_headers.append(
            str(header).strip()
        )

    # ------------------------------------------------------
    # DANH SÁCH KHOA
    # ------------------------------------------------------

    departments = []

    for row in range(
        106,
        ws_source.max_row + 1
    ):

        dept = ws_source.cell(
            row,
            4
        ).value

        if dept is None:

            continue

        dept = str(
            dept
        ).strip()

        if (
            dept.upper()
            == "TỔNG CỘNG"
        ):

            break

        if dept:

            departments.append(
                dept
            )

    # ------------------------------------------------------
    # DỮ LIỆU PHẦN MỀM
    # ------------------------------------------------------

    data = {
        dept: [None] * len(
            software_columns
        )
        for dept in departments
    }

    for row in range(
        106,
        ws_source.max_row + 1
    ):

        dept = ws_source.cell(
            row,
            4
        ).value

        if dept is None:

            continue

        dept = str(
            dept
        ).strip()

        if (
            dept.upper()
            == "TỔNG CỘNG"
        ):

            break

        if dept not in data:

            continue

        for i, col in enumerate(
            software_columns
        ):

            value = ws_source.cell(
                row,
                col
            ).value

            if value in (
                None,
                "",
                0
            ):

                data[
                    dept
                ][i] = None

            else:

                data[
                    dept
                ][i] = value

    # ------------------------------------------------------
    # SỬA LỖI KHÁC
    # B7:B105 = Khoa
    # AI7:AI105 = Sửa chữa khác
    # ------------------------------------------------------

    other_errors = {
        dept: []
        for dept in departments
    }

    all_error_occurrences = []

    for row in range(
        7,
        106
    ):

        dept = ws_source.cell(
            row,
            2
        ).value

        error = ws_source.cell(
            row,
            35
        ).value

        if (
            dept is None
            or error is None
        ):

            continue

        dept = str(
            dept
        ).strip()

        error = str(
            error
        ).strip()

        if (
            dept in other_errors
            and error
            and error != "0"
        ):

            other_errors[
                dept
            ].append(
                error
            )

            all_error_occurrences.append(
                error
            )

    # ------------------------------------------------------
    # XÓA LỖI TRÙNG KHI HIỂN THỊ
    # ------------------------------------------------------

    for dept in other_errors:

        unique_errors = []

        seen = set()

        for error in other_errors[
            dept
        ]:

            if error not in seen:

                seen.add(
                    error
                )

                unique_errors.append(
                    error
                )

        other_errors[
            dept
        ] = unique_errors

    # ------------------------------------------------------
    # CHỈ GIỮ KHOA CÓ DỮ LIỆU
    # ------------------------------------------------------

    valid_departments = []

    for dept in departments:

        has_software_data = any(
            value is not None
            and value != ""
            and value != 0
            for value in data[
                dept
            ]
        )

        has_other_error = any(
            error is not None
            and str(error).strip() != ""
            and str(error).strip() != "0"
            for error in other_errors[
                dept
            ]
        )

        if (
            has_software_data
            or has_other_error
        ):

            valid_departments.append(
                dept
            )

    # ------------------------------------------------------
    # XÁC ĐỊNH CỘT CÓ DỮ LIỆU
    # ------------------------------------------------------

    active_software_indexes = []

    for i in range(
        len(software_columns)
    ):

        has_data = any(
            data[dept][i] is not None
            and data[dept][i] != ""
            and data[dept][i] != 0
            for dept in valid_departments
        )

        if has_data:

            active_software_indexes.append(
                i
            )

    has_other_error_column = any(
        other_errors[dept]
        for dept in valid_departments
    )

    # ------------------------------------------------------
    # CỘT XUẤT
    # ------------------------------------------------------

    output_columns = [
        ("stt", None),
        ("dept", None)
    ]

    for i in active_software_indexes:

        output_columns.append(
            ("software", i)
        )

    if has_other_error_column:

        output_columns.append(
            ("other_error", None)
        )

    total_columns = len(
        output_columns
    )

    # ======================================================
    # TẠO WORKBOOK
    # ======================================================

    result = Workbook()

    ws = result.active

    ws.title = (
        "BÁO CÁO TỔNG HỢP"
    )

    # ======================================================
    # TIÊU ĐỀ
    # ======================================================

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=total_columns
    )

    title_cell = ws.cell(
        1,
        1
    )

    title_cell.value = (
        "TỔNG HỢP CÔNG VIỆC P.CNTT"
    )

    title_cell.font = Font(
        name="Arial",
        size=16,
        bold=True
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[
        1
    ].height = 32

    # ======================================================
    # HEADER
    # ======================================================

    headers = [
        "STT",
        "KHOA, PHÒNG, TRUNG TÂM"
    ]

    for i in active_software_indexes:

        headers.append(
            software_headers[i]
        )

    if has_other_error_column:

        headers.append(
            "SỬA LỖI KHÁC"
        )

    thin = Side(
        style="thin"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    header_row = 3

    data_start_row = 4

    for col, header in enumerate(
        headers,
        1
    ):

        cell = ws.cell(
            header_row,
            col
        )

        cell.value = header

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True
        )

        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[
        header_row
    ].height = 75

    # ======================================================
    # GHI DỮ LIỆU
    # ======================================================

    for index, dept in enumerate(
        valid_departments,
        1
    ):

        row = (
            data_start_row
            + index
            - 1
        )

        output_col = 1

        # STT
        ws.cell(
            row,
            output_col,
            index
        )

        output_col += 1

        # KHOA
        ws.cell(
            row,
            output_col,
            dept
        )

        output_col += 1

        # PHẦN MỀM
        for i in active_software_indexes:

            value = data[
                dept
            ][i]

            if value not in (
                None,
                "",
                0
            ):

                ws.cell(
                    row,
                    output_col,
                    value
                )

            output_col += 1

        # SỬA LỖI KHÁC
        if has_other_error_column:

            if other_errors[
                dept
            ]:

                ws.cell(
                    row,
                    output_col,
                    ", ".join(
                        other_errors[
                            dept
                        ]
                    )
                )

            output_col += 1

        # ĐỊNH DẠNG
        for col in range(
            1,
            total_columns + 1
        ):

            cell = ws.cell(
                row,
                col
            )

            cell.font = Font(
                name="Arial",
                size=10
            )

            cell.border = border

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if col in (
                        2,
                        total_columns
                    )
                    else "center"
                ),
                vertical="top",
                wrap_text=True
            )

    # ======================================================
    # TỔNG CỘNG
    # ======================================================

    total_row = (
        data_start_row
        + len(valid_departments)
    )

    ws.cell(
        total_row,
        2,
        "TỔNG CỘNG"
    )

    # ------------------------------------------------------
    # CỘNG CÁC CỘT SỐ
    # ------------------------------------------------------

    for col in range(
        3,
        total_columns + 1
    ):

        header = ws.cell(
            header_row,
            col
        ).value

        if (
            header
            == "SỬA LỖI KHÁC"
        ):

            continue

        values = []

        for row in range(
            data_start_row,
            total_row
        ):

            value = ws.cell(
                row,
                col
            ).value

            if isinstance(
                value,
                (
                    int,
                    float
                )
            ):

                values.append(
                    value
                )

        if values:

            total_value = sum(
                values
            )

            if total_value != 0:

                ws.cell(
                    total_row,
                    col,
                    total_value
                )

    # ======================================================
    # TỔNG SỬA LỖI KHÁC
    # ======================================================

    if has_other_error_column:

        other_error_total = 0

        for col in range(
            3,
            total_columns
        ):

            header = str(
                ws.cell(
                    header_row,
                    col
                ).value
                or ""
            ).strip()

            if (
                header
                == "Đăng bài trên website"
            ):

                continue

            value = ws.cell(
                total_row,
                col
            ).value

            if isinstance(
                value,
                (
                    int,
                    float
                )
            ):

                other_error_total += value

        # Không tính Đăng bài trên website
        for col in range(
            3,
            total_columns
        ):

            header = str(
                ws.cell(
                    header_row,
                    col
                ).value
                or ""
            ).strip()

            if (
                header
                == "Đăng bài trên website"
            ):

                ws.cell(
                    total_row,
                    col
                ).value = None

                break

        if (
            other_error_total
            != 0
        ):

            ws.cell(
                total_row,
                total_columns,
                other_error_total
            )

    # ======================================================
    # ĐỊNH DẠNG TỔNG CỘNG
    # ======================================================

    for col in range(
        1,
        total_columns + 1
    ):

        cell = ws.cell(
            total_row,
            col
        )

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True
        )

        cell.border = border

        cell.alignment = Alignment(
            horizontal=(
                "left"
                if col in (
                    2,
                    total_columns
                )
                else "center"
            ),
            vertical="top",
            wrap_text=True
        )

    # ======================================================
    # ĐỘ RỘNG CỘT
    # ======================================================

    ws.column_dimensions[
        "A"
    ].width = 7

    ws.column_dimensions[
        "B"
    ].width = 28

    for col in range(
        3,
        total_columns + 1
    ):

        header = ws.cell(
            header_row,
            col
        ).value

        if (
            header
            == "SỬA LỖI KHÁC"
        ):

            ws.column_dimensions[
                get_column_letter(col)
            ].width = 55

        else:

            ws.column_dimensions[
                get_column_letter(col)
            ].width = 17

    ws.freeze_panes = "C4"

    ws.auto_filter.ref = (
        f"A3:"
        f"{get_column_letter(total_columns)}"
        f"{total_row}"
    )

    # ======================================================
    # LƯU FILE
    # ======================================================

    result.save(
        output_file
    )

    return output_file


# ==========================================================
# ==========================================================
# GIAO DIỆN CHÍNH
# ==========================================================
# ==========================================================


class MainApplication:

    def __init__(
        self,
        root
    ):

        self.root = root

        self.root.title(
            "Báo cáo cổng giám định"
        )

        self.root.geometry(
            "650x480"
        )

        self.root.resizable(
            False,
            False
        )

        self.create_widgets()

    # ======================================================
    # GIAO DIỆN
    # ======================================================

    def create_widgets(self):

        # --------------------------------------------------
        # TIÊU ĐỀ
        # --------------------------------------------------

        tk.Label(
            self.root,
            text="BÁO CÁO CỔNG GIÁM ĐỊNH",
            font=(
                "Arial",
                18,
                "bold"
            )
        ).pack(
            pady=25
        )

        tk.Label(
            self.root,
            text=(
                "CHỌN CHỨC NĂNG CẦN THỰC HIỆN"
            ),
            font=(
                "Arial",
                11
            )
        ).pack(
            pady=5
        )

        # --------------------------------------------------
        # KHUNG NÚT
        # --------------------------------------------------

        frame = tk.Frame(
            self.root
        )

        frame.pack(
            pady=30
        )

        # --------------------------------------------------
        # NÚT GIÁM ĐỊNH
        # --------------------------------------------------

        tk.Button(
            frame,
            text=(
                "BÁO CÁO\n"
                "CỔNG GIÁM ĐỊNH"
            ),
            command=(
                self.run_giam_dinh
            ),
            font=(
                "Arial",
                12,
                "bold"
            ),
            width=28,
            height=4
        ).grid(
            row=0,
            column=0,
            padx=15
        )

        # --------------------------------------------------
        # NÚT P.CNTT
        # --------------------------------------------------

        tk.Button(
            frame,
            text=(
                "BÁO CÁO CÔNG VIỆC\n"
                "P.CNTT"
            ),
            command=(
                self.run_cntt
            ),
            font=(
                "Arial",
                12,
                "bold"
            ),
            width=28,
            height=4
        ).grid(
            row=0,
            column=1,
            padx=15
        )

        # --------------------------------------------------
        # TRẠNG THÁI
        # --------------------------------------------------

        self.status_label = tk.Label(
            self.root,
            text=(
                "Sẵn sàng"
            ),
            fg="gray",
            font=(
                "Arial",
                10
            )
        )

        self.status_label.pack(
            pady=20
        )

        # --------------------------------------------------
        # THÔNG TIN
        # --------------------------------------------------

        tk.Label(
            self.root,
            text=(
                "Báo cáo cổng giám định\n"
                "→ DANH SÁCH THEO DÕI GIÁM ĐỊNH BẢO HIỂM.xlsx\n\n"
                "Báo cáo công việc P.CNTT\n"
                "→ TỔNG HỢP CÔNG VIỆC P.CNTT.xlsx"
            ),
            font=(
                "Arial",
                9
            ),
            fg="gray",
            justify="center"
        ).pack(
            pady=10
        )

    # ======================================================
    # CHẠY BÁO CÁO GIÁM ĐỊNH
    # ======================================================

    def run_giam_dinh(self):

        try:

            input_file = (
                filedialog.askopenfilename(
                    title=(
                        "Chọn file Excel "
                        "báo cáo cổng giám định"
                    ),
                    filetypes=[
                        (
                            "Excel files",
                            "*.xlsx *.xlsm"
                        ),
                        (
                            "All files",
                            "*.*"
                        )
                    ]
                )
            )

            if not input_file:

                return

            selected = (
                choose_day_range()
            )

            if not selected:

                return

            if (
                selected
                == "01-14"
            ):

                start_day = 1
                end_day = 14

            else:

                start_day = 15
                end_day = 31

            self.status_label.config(
                text=(
                    "Đang xử lý "
                    "báo cáo cổng giám định..."
                ),
                fg="blue"
            )

            self.root.update()

            output_file = (
                process_giam_dinh(
                    input_file,
                    start_day,
                    end_day
                )
            )

            if output_file:

                self.status_label.config(
                    text=(
                        "Đã hoàn thành "
                        "báo cáo cổng giám định"
                    ),
                    fg="green"
                )

                messagebox.showinfo(
                    "HOÀN THÀNH",
                    (
                        "Đã tạo báo cáo thành công!\n\n"
                        f"Tên file:\n"
                        f"{TEN_FILE_GIAM_DINH}\n\n"
                        f"Đường dẫn:\n"
                        f"{output_file}"
                    )
                )

        except Exception as e:

            self.status_label.config(
                text="Có lỗi",
                fg="red"
            )

            messagebox.showerror(
                "LỖI",
                (
                    "Không thể tạo báo cáo "
                    "cổng giám định:\n\n"
                    f"{e}"
                )
            )

    # ======================================================
    # CHẠY BÁO CÁO P.CNTT
    # ======================================================

    def run_cntt(self):

        try:

            input_file = (
                filedialog.askopenfilename(
                    title=(
                        "Chọn file Excel "
                        "báo cáo công việc P.CNTT"
                    ),
                    filetypes=[
                        (
                            "Excel files",
                            "*.xlsx *.xlsm"
                        ),
                        (
                            "All files",
                            "*.*"
                        )
                    ]
                )
            )

            if not input_file:

                return

            folder = os.path.dirname(
                input_file
            )

            output_file = os.path.join(
                folder,
                TEN_FILE_CNTT
            )

            if os.path.exists(
                output_file
            ):

                answer = (
                    messagebox.askyesno(
                        "File đã tồn tại",
                        (
                            f"File:\n"
                            f"{TEN_FILE_CNTT}\n\n"
                            "đã tồn tại.\n\n"
                            "Bạn có muốn ghi đè không?"
                        )
                    )
                )

                if not answer:

                    return

            self.status_label.config(
                text=(
                    "Đang xử lý "
                    "báo cáo công việc P.CNTT..."
                ),
                fg="blue"
            )

            self.root.update()

            result = process_cntt(
                input_file,
                output_file
            )

            if result:

                self.status_label.config(
                    text=(
                        "Đã hoàn thành "
                        "báo cáo công việc P.CNTT"
                    ),
                    fg="green"
                )

                messagebox.showinfo(
                    "HOÀN THÀNH",
                    (
                        "Đã tạo báo cáo thành công!\n\n"
                        f"Tên file:\n"
                        f"{TEN_FILE_CNTT}\n\n"
                        f"Đường dẫn:\n"
                        f"{result}"
                    )
                )

        except Exception as e:

            self.status_label.config(
                text="Có lỗi",
                fg="red"
            )

            messagebox.showerror(
                "LỖI",
                (
                    "Không thể tạo báo cáo "
                    "P.CNTT:\n\n"
                    f"{e}"
                )
            )


# ==========================================================
# CHẠY CHƯƠNG TRÌNH
# ==========================================================

if __name__ == "__main__":

    root = tk.Tk()

    app = MainApplication(
        root
    )

    root.mainloop()
